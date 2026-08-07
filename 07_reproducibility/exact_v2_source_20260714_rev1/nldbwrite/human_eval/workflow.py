import argparse
import csv
import hashlib
import json
import random
import sqlite3
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from nldbwrite.common import find_db_path, iter_jsonl, load_json, read_id_file, save_json, sha256_file, write_jsonl
from nldbwrite.eval.evaluate import execute_sqls, memory_copy


GOLD_JUDGE_FIELDS = [
    'judge_input_clarity', 'judge_operation_correctness', 'judge_table_correctness',
    'judge_column_correctness', 'judge_value_correctness', 'judge_gold_sql_effect',
    'judge_keep_fix_remove', 'gold_error_type',
]
OUTPUT_JUDGE_FIELDS = ['human_output_label', 'human_error_type', 'human_evaluator_agreement']
FACT_JUDGE_FIELDS = [
    'judge_fact_coverage',
    'judge_attribute_correctness',
    'judge_value_correctness',
    'judge_grouping_correctness',
    'judge_hallucinated_fact',
    'judge_missing_required_fact',
    'judge_keep_fix_remove',
    'fact_error_type',
]

DEFAULT_V5_OUTPUT_AUDIT_RUNS = [
    'qwen7b_m2_builder_full',
    'qwen7b_m2_cbr_hybrid_repair_k3',
    'qwen7b_m5_fact_first',
    'qwen7b_m5_facts_cbr_hybrid_k3',
]

LABELS = {
    'judge_input_clarity': ['clear', 'minor_ambiguous', 'ambiguous', 'invalid'],
    'judge_operation_correctness': ['correct', 'partial', 'wrong', 'unclear'],
    'judge_table_correctness': ['correct', 'partial', 'wrong', 'unclear'],
    'judge_column_correctness': ['correct', 'partial', 'wrong', 'unclear'],
    'judge_value_correctness': ['correct', 'partial', 'wrong', 'unclear'],
    'judge_gold_sql_effect': ['correct', 'equivalent', 'partial', 'wrong', 'unsafe', 'unclear'],
    'judge_keep_fix_remove': ['keep', 'fix', 'remove', 'needs_adjudication'],
    'gold_error_type': ['none', 'ambiguous_input', 'wrong_operation', 'wrong_table', 'wrong_column', 'wrong_value', 'missing_required_column', 'missing_row', 'extra_row', 'wrong_upsert_conflict', 'unsafe_sql', 'schema_context_insufficient', 'evaluator_issue', 'other'],
    'human_output_label': ['correct', 'partially_correct', 'wrong', 'unsafe', 'unclear_due_to_gold'],
    'human_error_type': ['none', 'invalid_json', 'wrong_operation', 'wrong_table', 'wrong_column', 'missing_required_column', 'missing_optional_column', 'wrong_value', 'wrong_row_count', 'duplicate_row', 'wrong_conflict_target', 'wrong_update_mask', 'foreign_key_dependency_error', 'constraint_violation', 'builder_failure', 'execution_failure', 'unsafe_sql', 'ambiguous_input', 'gold_error', 'evaluator_error', 'other'],
    'human_evaluator_agreement': ['agree', 'disagree_auto_too_strict', 'disagree_auto_too_lenient', 'cannot_decide'],
    'judge_fact_coverage': ['complete', 'partial', 'missing', 'unclear'],
    'judge_attribute_correctness': ['correct', 'partial', 'wrong', 'unclear'],
    'judge_grouping_correctness': ['correct', 'partial', 'wrong', 'unclear'],
    'judge_hallucinated_fact': ['no', 'yes', 'unclear'],
    'judge_missing_required_fact': ['no', 'yes', 'unclear'],
    'fact_error_type': ['none', 'missing_value', 'wrong_value', 'wrong_attribute', 'wrong_grouping', 'hallucinated_fact', 'missing_conflict_key', 'missing_required_value', 'other'],
    'adjudication_decision': ['keep_original', 'fix_gold', 'remove_sample', 'fix_evaluator', 'mark_ambiguous', 'mark_safe_but_partial', 'mark_unsafe', 'needs_rerun'],
}


def stable_hash(value: Any) -> str:
    return hashlib.sha256(json.dumps(value, ensure_ascii=False, sort_keys=True).encode('utf-8')).hexdigest()


def write_csv(rows: list[dict[str, Any]], path: Path, fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = fields or list(rows[0]) if rows else (fields or ['sample_id'])
    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)


def read_table(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if path.suffix.lower() == '.xlsx':
        from openpyxl import load_workbook
        ws = load_workbook(path, data_only=False).active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(x or '') for x in rows[0]]
        return [{headers[i]: (row[i] if i < len(row) and row[i] is not None else '') for i in range(len(headers))} for row in rows[1:]]
    with open(path, encoding='utf-8-sig', newline='') as f:
        return list(csv.DictReader(f))


def schema_column_count(profile: dict[str, Any]) -> int:
    return sum(len(table.get('columns') or []) for table in profile.get('tables') or [])


def schema_bucket(count: int) -> str:
    if count < 30:
        return 'small'
    if count <= 100:
        return 'medium'
    if count <= 300:
        return 'large'
    return 'very_large'


def sample_stratum(sample: dict[str, Any], profile: dict[str, Any]) -> str:
    op = str(sample.get('operation_type') or 'insert').lower()
    scope = str(sample.get('impact_scope') or '')
    rows = int(sample.get('row_count') or sample.get('num_records') or 1)
    text = str(sample.get('input_text') or '').lower()
    if str(sample.get('input_type')) == 'noisy_mixed':
        return 'noisy'
    if str(sample.get('auto_difficulty')) == 'hard' or 'ambiguous' in text:
        return 'hard_ambiguous'
    if op != 'insert':
        tables = {r.get('table') for r in sample.get('gold_records') or []}
        composite = any(len(t.get('primary_keys') or []) > 1 for t in profile.get('tables') or [] if t.get('name') in tables)
        return 'composite_upsert' if composite else 'simple_upsert'
    if 'relational' in scope or int(sample.get('table_count') or sample.get('num_tables') or 1) > 1:
        return 'relational'
    if rows >= 10:
        return 'bulk_insert'
    if rows > 1:
        return 'batch_insert'
    return 'single_insert'


def freeze_manifest(data_path: Path, split_path: Path, profile_dir: Path, results_root: Path, out: Path) -> dict[str, Any]:
    predictions = {}
    for run in sorted(results_root.glob('*')):
        if not run.is_dir():
            continue
        files = {}
        for name in ['raw_generations.jsonl', 'parsed_outputs.jsonl', 'pred_sql.jsonl', 'evaluation.jsonl', 'run_manifest.json']:
            p = run / name
            if p.exists():
                files[name] = sha256_file(p)
        if files:
            predictions[run.name] = files
    profiles = {p.name: sha256_file(p) for p in sorted(profile_dir.glob('*.json'))}
    manifest = {
        'dataset_version': 'augmented900_v2_final',
        'split_version': 'locked_test_668',
        'evaluator_version': 'normalized_full_user_table_state_v3',
        'sample_file': str(data_path),
        'sample_file_hash': sha256_file(data_path),
        'split_file': str(split_path),
        'split_hash': sha256_file(split_path),
        'selected_sample_ids_hash': sha256_file(split_path),
        'db_profile_hash': stable_hash(profiles),
        'profile_hashes': profiles,
        'prediction_files': predictions,
        'created_at': time.strftime('%Y-%m-%dT%H:%M:%S%z'),
        'notes': 'Human evaluation uses frozen data and frozen model outputs. Automatic state evaluation remains the primary metric.',
    }
    save_json(manifest, out)
    return manifest


def gold_sample_plan(data_path: Path, split_path: Path, profile_dir: Path, out: Path, n: int = 300, seed: int = 2026) -> list[dict[str, Any]]:
    allowed = read_id_file(split_path)
    data = [x for x in load_json(data_path) if str(x['id']) in allowed]
    profiles = {p.stem: load_json(p) for p in profile_dir.glob('*.json')}
    quotas = {
        'composite_upsert': 40, 'noisy': 25, 'hard_ambiguous': 25, 'relational': 50,
        'simple_upsert': 40, 'bulk_insert': 40, 'batch_insert': 40, 'single_insert': 40,
    }
    rng = random.Random(seed)
    groups = defaultdict(list)
    for sample in data:
        profile = profiles[sample['db_id']]
        op = str(sample.get('operation_type') or 'insert').lower()
        row_count = int(sample.get('row_count') or sample.get('num_records') or 1)
        table_count = int(sample.get('table_count') or sample.get('num_tables') or 1)
        composite = any(len(table.get('primary_keys') or []) > 1 for table in profile.get('tables') or [] if table.get('name') in set(sample.get('gold_tables') or []))
        eligible = []
        if op == 'insert' and row_count == 1:
            eligible.append('single_insert')
        if op == 'insert' and 1 < row_count < 10:
            eligible.append('batch_insert')
        if op == 'insert' and row_count >= 10:
            eligible.append('bulk_insert')
        if table_count > 1:
            eligible.append('relational')
        if op != 'insert':
            eligible.append('composite_upsert' if composite else 'simple_upsert')
        if str(sample.get('input_type')) == 'noisy_mixed':
            eligible.append('noisy')
        if str(sample.get('auto_difficulty')) == 'hard' or 'ambiguous' in str(sample.get('input_text') or '').lower():
            eligible.append('hard_ambiguous')
        for stratum in eligible or [sample_stratum(sample, profile)]:
            groups[stratum].append(sample)
    selected = []
    source_counts = Counter()
    selected_ids = set()
    for stratum, quota in quotas.items():
        candidates = list(groups[stratum])
        rng.shuffle(candidates)
        for sample in candidates:
            source = str(sample.get('source_group_id') or sample['id'])
            if str(sample['id']) in selected_ids or source_counts[source] >= 2:
                continue
            selected.append((sample, stratum))
            source_counts[source] += 1
            selected_ids.add(str(sample['id']))
            if sum(s == stratum for _, s in selected) >= quota:
                break
    remaining = [x for x in data if str(x['id']) not in selected_ids and source_counts[str(x.get('source_group_id') or x['id'])] < 2]
    rng.shuffle(remaining)
    for sample in remaining:
        if len(selected) >= n:
            break
        source = str(sample.get('source_group_id') or sample['id'])
        if source_counts[source] >= 2:
            continue
        selected.append((sample, sample_stratum(sample, profiles[sample['db_id']])))
        source_counts[source] += 1
    rows = []
    for sample, stratum in selected[:n]:
        profile = profiles[sample['db_id']]
        count = schema_column_count(profile)
        rows.append({
            'sample_id': sample['id'], 'source_group_id': sample.get('source_group_id'),
            'db_id': sample['db_id'], 'split': 'locked_test_668',
            'operation_type': sample.get('operation_type'), 'write_scope': sample.get('impact_scope'),
            'input_format': sample.get('input_type'), 'difficulty': sample.get('auto_difficulty'),
            'schema_size_bucket': schema_bucket(count), 'schema_column_count': count,
            'is_original': not bool(sample.get('is_augmented')), 'is_augmented': bool(sample.get('is_augmented')),
            'sampling_stratum': stratum, 'selected_for_gold_eval': True,
        })
    write_csv(rows, out)
    return rows


def relevant_schema(profile: dict[str, Any], sample: dict[str, Any]) -> str:
    target = set(sample.get('gold_tables') or [])
    lines = []
    for table in profile.get('tables') or []:
        if table.get('name') not in target:
            continue
        lines.append(f"Table: {table['name']}")
        lines.append(f"Primary key: {', '.join(table.get('primary_keys') or []) or '-'}")
        lines.append('Required: ' + (', '.join(table.get('required_insert_columns') or []) or '-'))
        unique = ['+'.join(x.get('columns') or []) for x in table.get('unique_indexes') or []]
        lines.append('Unique indexes: ' + (', '.join(unique) or '-'))
        for col in table.get('columns') or []:
            fk = col.get('foreign_key') or {}
            suffix = f" FK->{fk.get('to_table')}.{fk.get('to_column')}" if fk else ''
            lines.append(f"- {col['name']} {col.get('type') or ''}{' NOT NULL' if col.get('not_null') else ''}{suffix}")
        lines.append('')
    return '\n'.join(lines).strip()


def select_matching_rows(conn: sqlite3.Connection, records: list[dict[str, Any]], limit: int = 20) -> dict[str, list[dict[str, Any]]]:
    output = defaultdict(list)
    for record in records:
        table = record.get('table')
        values = record.get('values') or {}
        if not table or not values:
            continue
        clauses, params = [], []
        for col, value in values.items():
            if value is None:
                clauses.append(f'"{col}" IS NULL')
            else:
                clauses.append(f'"{col}" = ?')
                params.append(value)
        try:
            cur = conn.execute(f'SELECT * FROM "{table}" WHERE ' + ' AND '.join(clauses) + f' LIMIT {int(limit)}', params)
            cols = [d[0] for d in cur.description]
            output[table].extend(dict(zip(cols, row)) for row in cur.fetchall())
        except sqlite3.Error as exc:
            output[table].append({'context_error': str(exc)})
    return dict(output)


def gold_context(data_path: Path, plan_path: Path, profile_dir: Path, db_root: Path | None, out: Path) -> list[dict[str, Any]]:
    data = {str(x['id']): x for x in load_json(data_path)}
    profiles = {p.stem: load_json(p) for p in profile_dir.glob('*.json')}
    contexts = []
    for plan in read_table(plan_path):
        sample = data[str(plan['sample_id'])]
        initial = {}
        post = {'gold_records': sample.get('gold_records') or []}
        execution = {'available': False}
        if db_root and db_root.exists():
            conn = memory_copy(find_db_path(db_root, sample['db_id']))
            initial = select_matching_rows(conn, sample.get('gold_records') or [])
            ok, error, executed = execute_sqls(conn, sample.get('gold_sql') or [])
            post = select_matching_rows(conn, sample.get('gold_records') or []) if ok else {}
            execution = {'available': True, 'success': ok, 'error': error, 'executed': executed}
            conn.close()
        contexts.append({
            **plan, 'input_text': sample.get('input_text'), 'relevant_schema': relevant_schema(profiles[sample['db_id']], sample),
            'initial_relevant_rows': initial, 'gold_records': sample.get('gold_records') or [],
            'gold_sql': sample.get('gold_sql') or [], 'gold_post_state': post, 'gold_execution': execution,
        })
    write_jsonl(contexts, out)
    return contexts


def result_maps(results_root: Path) -> dict[str, dict[str, dict[str, Any]]]:
    runs = {}
    for run in sorted(results_root.glob('*')):
        if not (run / 'evaluation.jsonl').exists():
            continue
        runs[run.name] = {}
        for filename, key in [('evaluation.jsonl', 'eval'), ('parsed_outputs.jsonl', 'parsed'), ('pred_sql.jsonl', 'built')]:
            path = run / filename
            if path.exists():
                for row in iter_jsonl(path):
                    runs[run.name].setdefault(str(row['sample_id']), {})[key] = row
    return runs


def output_sample_plan(data_path: Path, results_root: Path, out: Path, n: int = 200, seed: int = 2026) -> list[dict[str, Any]]:
    data = {str(x['id']): x for x in load_json(data_path)}
    runs = result_maps(results_root)
    if len(runs) < 2:
        raise RuntimeError('Output audit requires at least two completed runs with evaluation.jsonl.')
    run_names = sorted(runs)
    candidates = defaultdict(list)
    common_ids = set.intersection(*(set(runs[name]) for name in run_names))
    for sid in common_ids:
        labels = {name: bool((runs[name][sid].get('eval') or {}).get('correct')) for name in run_names}
        sample = data[sid]
        if all(labels.values()):
            reason = 'all_correct'
        elif not any(labels.values()):
            reason = 'all_wrong'
        elif labels.get('qwen7b_m2_extract_builder') is False and any(v for k, v in labels.items() if 'cbr' in k or 'repair' in k):
            reason = 'm2_wrong_cbr_correct'
        elif labels.get('qwen7b_m2_extract_builder') is True and any(not v for k, v in labels.items() if 'cbr' in k or 'repair' in k):
            reason = 'm2_correct_cbr_wrong'
        elif str(sample.get('operation_type')).lower() != 'insert':
            reason = 'upsert_failure'
        elif int(sample.get('table_count') or 1) > 1:
            reason = 'relational_failure'
        elif int(sample.get('row_count') or 1) >= 10:
            reason = 'large_batch_failure'
        else:
            reason = 'method_disagreement'
        candidates[reason].append((sid, labels))
    quotas = {'all_correct': 25, 'all_wrong': 25, 'm2_wrong_cbr_correct': 40, 'm2_correct_cbr_wrong': 30, 'upsert_failure': 30, 'relational_failure': 30, 'large_batch_failure': 20}
    rng = random.Random(seed)
    chosen = []
    for reason, quota in quotas.items():
        rows = candidates[reason]
        rng.shuffle(rows)
        chosen.extend((sid, labels, reason) for sid, labels in rows[:quota])
    seen = {sid for sid, _, _ in chosen}
    rest = [(sid, labels, reason) for reason, rows in candidates.items() for sid, labels in rows if sid not in seen]
    rng.shuffle(rest)
    chosen.extend(rest[:max(0, n - len(chosen))])
    rows = []
    for sid, labels, reason in chosen[:n]:
        sample = data[sid]
        rows.append({
            'sample_id': sid, 'db_id': sample['db_id'], 'source_group_id': sample.get('source_group_id'),
            'operation_type': sample.get('operation_type'), 'difficulty': sample.get('auto_difficulty'),
            'input_format': sample.get('input_type'), 'selected_reason': reason,
            'methods_to_compare': ';'.join(run_names), 'auto_eval_pattern': json.dumps(labels, sort_keys=True),
        })
    write_csv(rows, out)
    return rows


def parse_run_names(value: str | None) -> list[str]:
    if not value:
        return list(DEFAULT_V5_OUTPUT_AUDIT_RUNS)
    return [item.strip() for item in value.split(',') if item.strip()]


def output_sample_plan_50x4(
    data_path: Path,
    results_root: Path,
    out: Path,
    run_names: list[str],
    n: int = 50,
    seed: int = 2026,
) -> list[dict[str, Any]]:
    data = {str(x['id']): x for x in load_json(data_path)}
    runs = result_maps(results_root)
    missing = [name for name in run_names if name not in runs]
    if missing:
        raise RuntimeError(f'Missing audit run(s): {", ".join(missing)}')
    if len(run_names) < 2:
        raise RuntimeError('Output audit requires at least two audit runs.')

    common_ids = set.intersection(*(set(runs[name]) for name in run_names))
    if not common_ids:
        raise RuntimeError('No common sample IDs across requested audit runs.')

    baseline = run_names[0]
    fact_runs = [name for name in run_names if 'm5' in name or 'fact' in name]
    cbr_runs = [name for name in run_names if 'cbr' in name or 'repair' in name]
    rng = random.Random(seed)
    candidates = defaultdict(list)
    all_rows = []
    for sid in common_ids:
        sample = data.get(sid)
        if not sample:
            continue
        labels = {name: bool((runs[name][sid].get('eval') or {}).get('correct')) for name in run_names}
        if all(labels.values()):
            reason = 'all_correct'
        elif not any(labels.values()):
            reason = 'all_wrong'
        elif not labels.get(baseline, False) and any(labels.get(name, False) for name in cbr_runs + fact_runs):
            reason = 'm2_wrong_retrieval_or_fact_correct'
        elif labels.get(baseline, False) and any(not labels.get(name, False) for name in fact_runs):
            reason = 'm2_correct_fact_wrong'
        elif str(sample.get('operation_type')).lower() != 'insert' or int(sample.get('table_count') or 1) > 1 or str(sample.get('auto_difficulty')) == 'hard':
            reason = 'upsert_relational_or_hard'
        else:
            reason = 'method_disagreement'
        item = (sid, labels, reason)
        candidates[reason].append(item)
        all_rows.append(item)

    quotas = {
        'all_correct': 10,
        'all_wrong': 10,
        'm2_wrong_retrieval_or_fact_correct': 10,
        'm2_correct_fact_wrong': 10,
        'upsert_relational_or_hard': 10,
    }
    chosen = []
    seen = set()
    for reason, quota in quotas.items():
        rows = list(candidates[reason])
        rng.shuffle(rows)
        for sid, labels, selected_reason in rows:
            if sid in seen:
                continue
            chosen.append((sid, labels, selected_reason))
            seen.add(sid)
            if sum(item_reason == reason for _, _, item_reason in chosen) >= quota:
                break

    rest = [row for row in all_rows if row[0] not in seen]
    rng.shuffle(rest)
    chosen.extend(rest[:max(0, n - len(chosen))])

    rows = []
    for sid, labels, reason in chosen[:n]:
        sample = data[sid]
        rows.append({
            'sample_id': sid, 'db_id': sample['db_id'], 'source_group_id': sample.get('source_group_id'),
            'operation_type': sample.get('operation_type'), 'difficulty': sample.get('auto_difficulty'),
            'input_format': sample.get('input_type'), 'selected_reason': reason,
            'methods_to_compare': ';'.join(run_names), 'auto_eval_pattern': json.dumps(labels, sort_keys=True),
        })
    write_csv(rows, out)
    return rows


def blind_mapping(run_names: list[str], seed: int = 2026) -> dict[str, str]:
    shuffled = list(run_names)
    random.Random(seed).shuffle(shuffled)
    return {f'System_{chr(65 + i)}': name for i, name in enumerate(shuffled)}


def output_context(
    data_path: Path,
    plan_path: Path,
    profile_dir: Path,
    results_root: Path,
    db_root: Path | None,
    out: Path,
    mapping_out: Path,
    run_names: list[str] | None = None,
) -> list[dict[str, Any]]:
    data = {str(x['id']): x for x in load_json(data_path)}
    profiles = {p.stem: load_json(p) for p in profile_dir.glob('*.json')}
    runs = result_maps(results_root)
    plans = read_table(plan_path)
    if run_names is None:
        planned = []
        for plan in plans:
            planned.extend(name for name in str(plan.get('methods_to_compare') or '').split(';') if name)
        run_names = sorted(set(planned)) if planned else sorted(runs)
    mapping = blind_mapping(run_names)
    save_json(mapping, mapping_out)
    reverse = {run: blind for blind, run in mapping.items()}
    contexts = []
    for plan in plans:
        sid = str(plan['sample_id'])
        sample = data[sid]
        gold_post = {'gold_records': sample.get('gold_records') or []}
        if db_root and db_root.exists():
            conn = memory_copy(find_db_path(db_root, sample['db_id']))
            ok, _, _ = execute_sqls(conn, sample.get('gold_sql') or [])
            if ok:
                gold_post = select_matching_rows(conn, sample.get('gold_records') or [])
            conn.close()
        for run_name in str(plan['methods_to_compare']).split(';'):
            item = runs.get(run_name, {}).get(sid, {})
            built = item.get('built') or {}
            public_plan = {k: v for k, v in plan.items() if k not in {'methods_to_compare', 'auto_eval_pattern', 'selected_reason'}}
            pred_post = {}
            if db_root and db_root.exists() and built.get('pred_sql'):
                conn = memory_copy(find_db_path(db_root, sample['db_id']))
                ok, error, _ = execute_sqls(conn, built.get('pred_sql') or [])
                pred_post = select_matching_rows(conn, sample.get('gold_records') or []) if ok else {'execution_error': error}
                conn.close()
            contexts.append({
                **public_plan, 'system_id_blind': reverse[run_name], 'input_text': sample.get('input_text'),
                'relevant_schema': relevant_schema(profiles[sample['db_id']], sample),
                'gold_records': sample.get('gold_records') or [], 'gold_post_state': gold_post,
                'predicted_json': (item.get('parsed') or {}).get('pred_json'),
                'built_sql': built.get('pred_sql') or [], 'execution_status': (item.get('eval') or {}).get('execution_success'),
                'pred_post_state': pred_post, 'auto_eval_label': (item.get('eval') or {}).get('correct'),
                'auto_error_type': (item.get('eval') or {}).get('error_type'),
            })
    write_jsonl(contexts, out)
    return contexts


def fact_bucket(value: float) -> str:
    if value >= 0.90:
        return '0.90-1.00'
    if value >= 0.70:
        return '0.70-0.90'
    if value >= 0.50:
        return '0.50-0.70'
    return '<0.50'


def fact_sample_plan(data_path: Path, fact_run_dir: Path, out: Path, n: int = 200, seed: int = 2026) -> list[dict[str, Any]]:
    data = {str(x['id']): x for x in load_json(data_path)}
    metrics_path = fact_run_dir / 'fact_eval_per_sample.csv'
    parsed_path = fact_run_dir / 'parsed_outputs.jsonl'
    if metrics_path.exists():
        source_rows = read_table(metrics_path)
    elif parsed_path.exists():
        source_rows = [{'sample_id': row['sample_id']} for row in iter_jsonl(parsed_path)]
    else:
        raise RuntimeError(f'Fact audit requires {metrics_path} or {parsed_path}.')
    groups = defaultdict(list)
    for row in source_rows:
        sid = str(row.get('sample_id'))
        if sid not in data:
            continue
        try:
            f1 = float(row.get('attribute_value_f1') or row.get('value_f1') or 0.0)
        except ValueError:
            f1 = 0.0
        groups[fact_bucket(f1)].append(row)
    quotas = {'0.90-1.00': 50, '0.70-0.90': 50, '0.50-0.70': 50, '<0.50': 50}
    rng = random.Random(seed)
    selected = []
    seen = set()
    for bucket, quota in quotas.items():
        rows = groups[bucket]
        rng.shuffle(rows)
        for row in rows:
            sid = str(row.get('sample_id'))
            if sid in seen:
                continue
            selected.append((row, bucket))
            seen.add(sid)
            if sum(b == bucket for _, b in selected) >= quota:
                break
    rest = [(row, bucket) for bucket, rows in groups.items() for row in rows if str(row.get('sample_id')) not in seen]
    rng.shuffle(rest)
    selected.extend(rest[:max(0, n - len(selected))])
    plan_rows = []
    for row, bucket in selected[:n]:
        sid = str(row.get('sample_id'))
        sample = data[sid]
        plan_rows.append({
            'sample_id': sid,
            'db_id': sample.get('db_id'),
            'source_group_id': sample.get('source_group_id'),
            'operation_type': sample.get('operation_type'),
            'difficulty': sample.get('auto_difficulty'),
            'input_format': sample.get('input_type'),
            'fact_f1_bucket': bucket,
            'value_f1': row.get('value_f1', ''),
            'attribute_value_f1': row.get('attribute_value_f1', ''),
            'hallucinated_fact_rate': row.get('hallucinated_fact_rate', ''),
            'row_count_accuracy': row.get('row_count_accuracy', ''),
            'selected_reason': 'fact_validation_bucket',
        })
    write_csv(plan_rows, out)
    return plan_rows


def fact_context(data_path: Path, plan_path: Path, fact_run_dir: Path, out: Path) -> list[dict[str, Any]]:
    data = {str(x['id']): x for x in load_json(data_path)}
    parsed = {}
    parsed_path = fact_run_dir / 'parsed_outputs.jsonl'
    if parsed_path.exists():
        parsed = {str(row['sample_id']): row for row in iter_jsonl(parsed_path)}
    metrics = {}
    metrics_path = fact_run_dir / 'fact_eval_per_sample.csv'
    if metrics_path.exists():
        metrics = {str(row['sample_id']): row for row in read_table(metrics_path)}
    contexts = []
    for plan in read_table(plan_path):
        sid = str(plan['sample_id'])
        sample = data[sid]
        pred = parsed.get(sid, {})
        predicted_facts = (pred.get('pred_facts') or pred.get('pred_json') or {}).get('facts') if isinstance(pred.get('pred_facts') or pred.get('pred_json'), dict) else []
        contexts.append({
            **plan,
            'input_text': sample.get('input_text'),
            'gold_records': sample.get('gold_records') or [],
            'predicted_facts': predicted_facts or [],
            'fact_eval_auto_metrics': metrics.get(sid, {}),
        })
    write_jsonl(contexts, out)
    return contexts


def judge_fields_for_task(task: str) -> list[str]:
    if task == 'gold':
        return GOLD_JUDGE_FIELDS
    if task == 'output':
        return OUTPUT_JUDGE_FIELDS
    if task == 'fact':
        return FACT_JUDGE_FIELDS
    raise ValueError(f'Unknown task: {task}')


def annotation_rows(contexts: list[dict[str, Any]], task: str, annotator: str, limit: int | None = None) -> list[dict[str, Any]]:
    rows = []
    judge_fields = judge_fields_for_task(task)
    for context in contexts[:limit]:
        row = {}
        for key, value in context.items():
            row[key] = json.dumps(value, ensure_ascii=False, indent=2) if isinstance(value, (dict, list)) else value
        for field in judge_fields:
            row[field] = ''
        row.update({'comment': '', 'annotator_id': annotator, 'annotation_time': ''})
        rows.append(row)
    return rows


def export_xlsx(rows: list[dict[str, Any]], path: Path, task: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = {'gold': 'Gold Validation', 'output': 'Output Audit', 'fact': 'Fact Validation'}.get(task, 'Annotation')
    fields = list(rows[0]) if rows else ['sample_id']
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, '') for field in fields])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(wrap_text=True)
    for idx, field in enumerate(fields, start=1):
        letter = ws.cell(1, idx).column_letter
        ws.column_dimensions[letter].width = 18 if field not in {'input_text', 'relevant_schema', 'gold_records', 'gold_sql', 'gold_post_state', 'predicted_json', 'predicted_facts', 'fact_eval_auto_metrics', 'built_sql', 'pred_post_state', 'comment'} else 42
        for cell in ws[letter][1:]:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        if field in LABELS and len(rows):
            validation = DataValidation(type='list', formula1='"' + ','.join(LABELS[field]) + '"', allow_blank=True)
            ws.add_data_validation(validation)
            validation.add(f'{letter}2:{letter}{len(rows)+1}')
    wb.save(path)


def prepare_sheets(context_path: Path, out_dir: Path, task: str, primary_n: int, overlap_n: int) -> None:
    contexts = list(iter_jsonl(context_path))
    primary = annotation_rows(contexts, task, 'A', primary_n)
    overlap = annotation_rows(contexts, task, 'B', overlap_n)
    export_xlsx(primary, out_dir / f'annotator_A_{task}_{primary_n}.xlsx', task)
    export_xlsx(overlap, out_dir / f'annotator_B_{task}_overlap_{overlap_n}.xlsx', task)
    write_csv(primary, out_dir / f'annotator_A_{task}_{primary_n}.csv')
    write_csv(overlap, out_dir / f'annotator_B_{task}_overlap_{overlap_n}.csv')


def cohen_kappa(a: list[str], b: list[str]) -> tuple[float, float, dict[str, dict[str, int]]]:
    pairs = [(str(x), str(y)) for x, y in zip(a, b) if str(x) and str(y)]
    if not pairs:
        return 0.0, 0.0, {}
    labels = sorted({x for pair in pairs for x in pair})
    matrix = {x: {y: 0 for y in labels} for x in labels}
    for x, y in pairs:
        matrix[x][y] += 1
    n = len(pairs)
    agreement = sum(x == y for x, y in pairs) / n
    pa = Counter(x for x, _ in pairs)
    pb = Counter(y for _, y in pairs)
    expected = sum((pa[label] / n) * (pb[label] / n) for label in labels)
    kappa = (agreement - expected) / (1 - expected) if expected < 1 else 1.0
    return agreement, kappa, matrix


def analyze_annotations(a_path: Path, b_path: Path, task: str, out_dir: Path) -> dict[str, Any]:
    a_rows, b_rows = read_table(a_path), read_table(b_path)
    key_fields = ['sample_id'] + (['system_id_blind'] if task == 'output' else [])
    def key(row): return tuple(str(row.get(field, '')) for field in key_fields)
    a_map, b_map = {key(row): row for row in a_rows}, {key(row): row for row in b_rows}
    common = sorted(set(a_map) & set(b_map))
    fields = judge_fields_for_task(task)
    reports, disagreements = [], []
    full_report = {'task': task, 'overlap': len(common), 'fields': {}}
    for field in fields:
        values_a = [str(a_map[k].get(field, '')) for k in common]
        values_b = [str(b_map[k].get(field, '')) for k in common]
        agreement, kappa, matrix = cohen_kappa(values_a, values_b)
        diff_count = sum(x != y for x, y in zip(values_a, values_b) if x and y)
        report = {'task': task, 'field': field, 'overlap': len(common), 'percent_agreement': agreement, 'cohen_kappa': kappa, 'disagreement_count': diff_count}
        reports.append(report)
        full_report['fields'][field] = {**report, 'confusion_matrix': matrix}
        for k, x, y in zip(common, values_a, values_b):
            if x != y or x in {'fix', 'remove', 'needs_adjudication'} or y in {'fix', 'remove', 'needs_adjudication'}:
                disagreements.append({
                    **{name: value for name, value in zip(key_fields, k)}, 'field_name': field,
                    'annotator_A_label': x, 'annotator_B_label': y,
                    'annotator_A_comment': a_map[k].get('comment', ''), 'annotator_B_comment': b_map[k].get('comment', ''),
                    'context_summary': str(a_map[k].get('input_text', ''))[:1500],
                    'adjudicated_label': '', 'adjudicated_error_type': '', 'adjudication_decision': '', 'adjudicator_comment': '',
                })
    out_dir.mkdir(parents=True, exist_ok=True)
    write_csv(reports, out_dir / f'agreement_report_{task}.csv')
    save_json(full_report, out_dir / f'agreement_report_{task}.json')
    write_csv(disagreements, out_dir / f'adjudication_{task}_queue.csv')
    export_xlsx(disagreements, out_dir / f'adjudication_{task}_queue.xlsx', task)
    return full_report


def make_corrections(adjudicated_path: Path, task: str, out_path: Path) -> None:
    corrections = []
    for row in read_table(adjudicated_path):
        decision = str(row.get('adjudication_decision') or '')
        if decision not in {'fix_gold', 'remove_sample', 'fix_evaluator', 'mark_ambiguous', 'mark_unsafe', 'needs_rerun'}:
            continue
        corrections.append({
            'sample_id': row.get('sample_id'), 'system_id_blind': row.get('system_id_blind'),
            'field_name': row.get('field_name'), 'issue_type': row.get('adjudicated_error_type') or row.get('adjudicated_label'),
            'decision': decision, 'reason': row.get('adjudicator_comment'), 'task': task,
        })
    write_jsonl(corrections, out_path)


def latex_table(rows: list[dict[str, Any]], fields: list[str], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\\begin{tabular}{' + 'l' + 'r' * (len(fields) - 1) + '}\n\\toprule\n')
        f.write(' & '.join(fields) + ' \\\\\n\\midrule\n')
        for row in rows:
            f.write(' & '.join(str(row.get(field, '-')) for field in fields) + ' \\\\\n')
        f.write('\\bottomrule\n\\end{tabular}\n')


def paper_tables(gold_final: Path | None, output_final: Path | None, fact_final: Path | None, agreement_dir: Path, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    agreement_rows = []
    for path in sorted(agreement_dir.glob('agreement_report_*.csv')):
        agreement_rows.extend(read_table(path))
    write_csv(agreement_rows, out_dir / 'human_agreement.csv')
    latex_table(agreement_rows, ['field', 'overlap', 'percent_agreement', 'cohen_kappa'], out_dir / 'human_agreement.tex')
    if gold_final and gold_final.exists():
        rows = read_table(gold_final)
        counts = Counter(str(row.get('judge_keep_fix_remove') or row.get('adjudicated_label') or 'unlabeled') for row in rows)
        summary = [{'Category': key, 'Reviewed': value} for key, value in sorted(counts.items())]
        write_csv(summary, out_dir / 'human_dataset_validation.csv')
        latex_table(summary, ['Category', 'Reviewed'], out_dir / 'human_dataset_validation.tex')
    if output_final and output_final.exists():
        rows = read_table(output_final)
        audit = Counter((str(row.get('auto_eval_label')), str(row.get('human_output_label') or row.get('adjudicated_label'))) for row in rows)
        audit_rows = [{'AutoEval': key[0], 'HumanEval': key[1], 'Count': count, 'Rate': count / len(rows) if rows else 0} for key, count in sorted(audit.items())]
        write_csv(audit_rows, out_dir / 'evaluator_audit.csv')
        latex_table(audit_rows, ['AutoEval', 'HumanEval', 'Count', 'Rate'], out_dir / 'evaluator_audit.tex')
        errors = Counter(str(row.get('human_error_type') or row.get('adjudicated_error_type') or 'none') for row in rows)
        error_rows = [{'ErrorType': key, 'Count': value} for key, value in sorted(errors.items())]
        write_csv(error_rows, out_dir / 'human_error_analysis.csv')
        latex_table(error_rows, ['ErrorType', 'Count'], out_dir / 'human_error_analysis.tex')
    if fact_final and fact_final.exists():
        rows = read_table(fact_final)
        counts = Counter(str(row.get('judge_keep_fix_remove') or row.get('adjudicated_label') or 'unlabeled') for row in rows)
        summary = [{'Category': key, 'Reviewed': value} for key, value in sorted(counts.items())]
        write_csv(summary, out_dir / 'human_fact_validation.csv')
        latex_table(summary, ['Category', 'Reviewed'], out_dir / 'human_fact_validation.tex')
        errors = Counter(str(row.get('fact_error_type') or row.get('adjudicated_error_type') or 'none') for row in rows)
        error_rows = [{'ErrorType': key, 'Count': value} for key, value in sorted(errors.items())]
        write_csv(error_rows, out_dir / 'human_fact_error_analysis.csv')
        latex_table(error_rows, ['ErrorType', 'Count'], out_dir / 'human_fact_error_analysis.tex')


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest='command', required=True)
    common = argparse.ArgumentParser(add_help=False)
    common.add_argument('--data', default='data/processed/nl_db_write_augmented900_v1.json')
    common.add_argument('--split', default='data/splits/augmented900_v1/test_ids.txt')
    common.add_argument('--profiles', default='artifacts/profiles_aug900')
    common.add_argument('--db-root', default='data/bird_databases')
    common.add_argument('--results-root', default='results/server_aug900_v2_final')
    common.add_argument('--fact-run-dir', default='results/server_aug900_v2_final/qwen7b_s_cbr_h_repair')
    common.add_argument('--root', default='human_eval')
    sub.add_parser('prepare-gold', parents=[common])
    prepare_output = sub.add_parser('prepare-output', parents=[common])
    prepare_output.add_argument('--output-design', choices=['legacy', 'v5-50x4'], default='legacy')
    prepare_output.add_argument('--audit-runs', default='')
    prepare_output.add_argument('--output-unique-samples', type=int, default=50)
    sub.add_parser('prepare-fact', parents=[common])
    analyze = sub.add_parser('analyze')
    analyze.add_argument('--annotator-a', required=True); analyze.add_argument('--annotator-b', required=True)
    analyze.add_argument('--task', choices=['gold', 'output', 'fact'], required=True); analyze.add_argument('--out-dir', default='human_eval/final')
    corr = sub.add_parser('corrections')
    corr.add_argument('--adjudicated', required=True); corr.add_argument('--task', choices=['gold', 'output', 'fact'], required=True); corr.add_argument('--out', required=True)
    tables = sub.add_parser('paper-tables')
    tables.add_argument('--gold-final'); tables.add_argument('--output-final'); tables.add_argument('--fact-final'); tables.add_argument('--agreement-dir', default='human_eval/final'); tables.add_argument('--out-dir', default='human_eval/paper_tables')
    args = ap.parse_args()
    if args.command in {'prepare-gold', 'prepare-output', 'prepare-fact'}:
        root = Path(args.root); artifacts = root / 'artifacts'; sheets = root / 'sheets'
        save_json({'labels': LABELS, 'gold_fields': GOLD_JUDGE_FIELDS, 'output_fields': OUTPUT_JUDGE_FIELDS, 'fact_fields': FACT_JUDGE_FIELDS}, root / 'config' / 'annotation_schema.json')
        data, split, profiles, results = Path(args.data), Path(args.split), Path(args.profiles), Path(args.results_root)
        db_root = Path(args.db_root)
        freeze_manifest(data, split, profiles, results, artifacts / 'frozen_manifest.json')
        if args.command == 'prepare-gold':
            plan = gold_sample_plan(data, split, profiles, artifacts / 'sample_plan_gold_300.csv')
            contexts = gold_context(data, artifacts / 'sample_plan_gold_300.csv', profiles, db_root if db_root.exists() else None, artifacts / 'gold_eval_context.jsonl')
            prepare_sheets(artifacts / 'gold_eval_context.jsonl', sheets, 'gold', min(300, len(contexts)), min(100, len(contexts)))
        else:
            if args.command == 'prepare-output':
                if args.output_design == 'v5-50x4':
                    audit_runs = parse_run_names(args.audit_runs)
                    plan_path = artifacts / 'sample_plan_output_50x4.csv'
                    plan = output_sample_plan_50x4(data, results, plan_path, audit_runs, n=args.output_unique_samples)
                    contexts = output_context(
                        data, plan_path, profiles, results, db_root if db_root.exists() else None,
                        artifacts / 'output_audit_context.jsonl', artifacts / 'blind_mapping.json', audit_runs,
                    )
                else:
                    plan_path = artifacts / 'sample_plan_output_200.csv'
                    plan = output_sample_plan(data, results, plan_path)
                    contexts = output_context(
                        data, plan_path, profiles, results, db_root if db_root.exists() else None,
                        artifacts / 'output_audit_context.jsonl', artifacts / 'blind_mapping.json',
                    )
                prepare_sheets(artifacts / 'output_audit_context.jsonl', sheets, 'output', min(200, len(contexts)), min(100, len(contexts)))
            else:
                fact_run = Path(args.fact_run_dir)
                plan = fact_sample_plan(data, fact_run, artifacts / 'sample_plan_fact_200.csv')
                contexts = fact_context(data, artifacts / 'sample_plan_fact_200.csv', fact_run, artifacts / 'fact_eval_context.jsonl')
                prepare_sheets(artifacts / 'fact_eval_context.jsonl', sheets, 'fact', min(200, len(contexts)), min(100, len(contexts)))
    elif args.command == 'analyze':
        analyze_annotations(Path(args.annotator_a), Path(args.annotator_b), args.task, Path(args.out_dir))
    elif args.command == 'corrections':
        make_corrections(Path(args.adjudicated), args.task, Path(args.out))
    else:
        paper_tables(Path(args.gold_final) if args.gold_final else None, Path(args.output_final) if args.output_final else None, Path(args.fact_final) if args.fact_final else None, Path(args.agreement_dir), Path(args.out_dir))


if __name__ == '__main__':
    main()
